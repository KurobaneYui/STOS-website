#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tqdm, json, time, random, os, datetime
import pandas as pd
from mysql import connector
from openpyxl import Workbook

class STOS_DB_conn:
    def __init__(self,set_code,set_config,set_order=(False,0)):
        '''
set_code：-p 指定设定文件的路径；-c 直接给定参数字典；-h 获取帮助信息
set_config：根据set_code的参数，提供路径或者字典
set_order：元组：第一个元素：bool型，为True时，使用第二个元素指定编号的设定；第二个元素：int型，第一个元素为True时，指定使用设定文件的第几个设置
        '''
        self.querys_list = [];
        self.results_list = [];
        if(set_code=="-p"):
            with open(set_config) as f:
                config = []
                a = f.readline()
                while(a):
                    config.append(json.loads(a))
                    a = f.readline()
                if(set_order[0]==False):
                    print("使用哪一个数据库对应的config设定？")
                    temp_count = 0
                    for conf in config:
                        print("\t"+str(temp_count)+") "+conf["database"])
                        temp_count += 1
                    order = int(input())
                    while(order>=temp_count):
                        order = input("输入数字不在范围，请重新输入：")
                elif(set_order[0]==True):
                    temp_count = 0
                    for conf in config:
                        temp_count += 1
                    if(set_order[1]>=temp_count):
                        raise "ERROR! 数据库链接配置信息不存在"
                    else:
                        order = set_order[1]
                else:
                    raise "ERROR! 错误的参数：不正确的配置信息参数"
                self.conn = connector.connect(**config[order])
                self.cur = self.conn.cursor()
        elif (set_code=="-c"):
            self.conn = connector.connect(**set_config)
            self.cur = self.conn.cursor()
        elif (set_code=="-h"):
            print("-p load config from json file which the path of is given")
            print("-c give the config")
            print("-h get help info")
        else:
            print("ERROR! unknown set_code: "+set_code)
            print()
            print("-p load config from json file which the path of is given")
            print("-c give the config")
            print("-h get help info")
            return (False, "ERROR! unknown set_code: "+set_code)

    def __del__(self): # 析构函数
        self.cur.close()
        self.conn.close()

    def execute_query(self, sql): # 直接运行sql语句，在有返回的条件下返回结果
        self.cur.execute(sql)
        try:
            result = self.cur.fetchall()
            return result
        except:
            self.conn.commit()
            return

    def push_query_list(self, sql): # 插入一个待执行的sql语句
        self.querys_list.append(sql)
        return True
    
    def commit_query_list(self): # 执行已输入的sql语句
        self.results_list = []
        for sql in self.querys_list:
            self.cur.execute(sql)
            try:
                result = self.cur.fetchall()
                self.results_list.append(result)
            except:
                self.results_list.append("")
        self.conn.commit()
        self.querys_list = []

    def get_queryList_results(self): # 获取sql执行后的结果
        ret = self.results_list
        self.results_list = []
        return ret

    def clean_query_list(self): # 清空待执行的sql语句
        self.querys_list = []

class zaozixi: # 早自习类，提供相关数据的上传和下载
    def __init__(self):
        self.con_info = STOS_DB_conn("-p", "./config.json", (True,0))
        self.con_data = STOS_DB_conn("-p", "./config.json", (True,1))

    def __del__(self):
        del self.con_info
        del self.con_data

    def work_info(self, path): # 导入教室数据
        work_info = pd.read_excel(path)

        print("数据标题为：")
        print(list(work_info.columns))
        if(input("是否使用此文件？（y/N）：")!="y"):
            return

        qvhao_jiaoshibianhao = work_info.iloc[:,2]
        xueyuan = work_info.iloc[:,1]
        yingdaorenshu = work_info.iloc[:,3]

        zhoushu = input("请输入导入数据为第几周，例如：第十一周\n")
        print("请输入该周周一对应的日期，例如：2019-05-06")
        while(1):
            try:
                time_start = datetime.datetime.strptime(input(),"%Y-%m-%d")
                break
            except:
                print("日期错误，请重新输入")
        print("已设定起始日期为：{}".format(time_start))

        sql = "INSERT INTO `查早排班`(`教学楼`, `区号`, `教室编号`, `学院`, `应到人数`, `周数`, `周起始日期`) VALUES ('品学楼','{}','{}','{}','{}','{}','{}');"
        for i in range(qvhao_jiaoshibianhao.count()):
            self.con_data.push_query_list(sql.format(qvhao_jiaoshibianhao.iloc[i][0], qvhao_jiaoshibianhao.iloc[i][1:], xueyuan.iloc[i], yingdaorenshu.iloc[i], zhoushu, time_start))
        self.con_data.commit_query_list()

    def work_schedule_manuel(self, path): # 手动导入查早排班
        work_info = pd.read_excel(path)

        print("数据标题为：")
        print(list(work_info.columns))
        if(input("是否使用此文件？（y/N）：")!="y"):
            return

        xuehao = work_info.iloc[:,0]
        xingming = work_info.iloc[:,1]
        jiaoxuelou = work_info.iloc[:,2]
        qvhao = work_info.iloc[:,3]
        jiaoshibianhao = work_info.iloc[:,4]

        print("请输入该周周一对应的日期，例如：2019-05-06")
        while(1):
            try:
                time_start = datetime.datetime.strptime(input(),"%Y-%m-%d")
                break
            except:
                print("日期错误，请重新输入")
        print("已设定起始日期为：{}".format(time_start))

        sql = "UPDATE `查早排班` SET `查早组员`='{}',`姓名`='{}' WHERE `教学楼`='{}' AND `区号`='{}' AND `教室编号`='{}' AND `周起始日期`='{}';"
        for i in range(jiaoshibianhao.count()):
            self.con_data.push_query_list(sql.format(xuehao.iloc[i], xingming.iloc[i], jiaoxuelou.iloc[i], qvhao.iloc[i], jiaoshibianhao.iloc[i], time_start))
        self.con_data.commit_query_list()

    def work_schedule_auto(self): # 自动导入查早排班
        print("请输入该周周一对应的日期，例如：2019-05-06")
        while(1):
            try:
                time_start = datetime.datetime.strptime(input(),"%Y-%m-%d")
                break
            except:
                print("日期错误，请重新输入")
        print("已设定起始日期为：{}".format(time_start))

        if os.path.exists('./group_sequence.json'):
            with open('./group_sequence.json','r') as f:
                group = json.load(f)
        else:
            group=['现场组一组','现场组二组','现场组三组','现场组四组','现场组五组','现场组六组']
        #调整顺序后保存
        new_group = [group[-1]]+group[0:5]
        with open('./group_sequence.json','w') as f:
            json.dump(new_group,f)

        '''
        这里把每组学号的列表构建成一个新的列表group_member
        [[1组员学号],
        [2组员学号],
        ...
        [6组员学号]]

        当要组间教室轮换，只要改变此列表内部组内顺序即可
        '''
        group_member=[]
        for g in group:
            sql = "SELECT 学号 FROM 成员岗位 WHERE 所属组='"+g+"' and 岗位='组员';"
            group_member.append([])
            for i in self.con_info.execute_query(sql):
                group_member[-1].append(i[0])
        group_num=[len(i) for i in group_member]

        '''
        实现这样的排序
        [('品学楼', 'A', 301, 82),
         ('品学楼', 'A', 103, 82),
         ('品学楼', 'A', 102, 50),
         ('品学楼', 'A', 101, 57),
         ('品学楼', 'B', 102, 82),
         ('品学楼', 'B', 103, 82),
         ('品学楼', 'B', 401, 90)]
        '''
        sql = "SELECT `教学楼`, `区号`, `教室编号` FROM `查早排班` WHERE `周起始日期`='{}';".format(time_start)
        class_room = self.con_data.execute_query(sql)
        temp=sorted(class_room,key=lambda x:x[2],reverse=True)
        class_room=sorted(temp,key=lambda x:x[1]=='B' and int(x[2]))

        ArrangeResult=[] #存放排班结果
        index=0
        for i in range(len(group_num)):
            class_part=class_room[index:index+group_num[i]]#选择各组对应的教室
            index += group_num[i]
            random.seed(time.time())
            random.shuffle(class_part)#打乱顺序
            for j in range(group_num[i]):#为组内队员排班
                temp=(group_member[i][j],)+class_part[j]
                ArrangeResult.append(temp)
        
        for i in ArrangeResult:
            sql = "SELECT `姓名` FROM `成员信息` WHERE `学号`="+i[0]+";"
            xingming = self.con_info.execute_query(sql)[0][0]

            sql = "UPDATE `查早排班` SET `查早组员`='{}',`姓名`='{}' WHERE `教学楼`='{}' AND `区号`='{}' AND `教室编号`='{}' AND `周起始日期`='{}';"
            self.con_data.push_query_list(sql.format(i[0],xingming,i[1],i[2],i[3],time_start))
        self.con_data.commit_query_list()

    def download(self, path): # 下载数据
        print("开始导出早自习缺勤表")
        self._download_queqin(path)
        print("导出早自习缺勤表完成")

        print("开始导出早自习人数表")
        self._download_renshu(path)
        print("导出早自习人数表完成")

    def _download_renshu(self, path): # 下载早自习教室人数数据
        path = os.path.join(path,"早自习人数数据.xlsx")
        print("请输入导出数据的起始日期，例如：2019-05-06")
        while(1):
            try:
                time_start = datetime.datetime.strptime(input(),"%Y-%m-%d")
                break
            except:
                print("日期错误，请重新输入")
        print("已设定起始日期为：{}".format(time_start))
        print("请输入导出数据的结束日期，例如：2019-05-06\n直接回车不输入日期则默认导出一周数据")
        while(1):
            a = input()
            if a=="":
                time_end = time_start+datetime.timedelta(days=6)
                break
            else:
                try:
                    time_end = datetime.datetime.strptime(a,"%Y-%m-%d")
                    break
                except:
                    print("日期错误，请重新输入")
        print("已设定结束日期为：{}".format(time_end))

        sql = "SELECT `日期`,`教学楼`,`区号`,`教室编号`,`教室数据`,`提交者` FROM `查早数据` WHERE `日期` BETWEEN '{}' AND '{}' ORDER BY `教学楼`,`区号`,`教室编号`,`日期` ASC;".format(time_start,time_end)
        results = self.con_data.execute_query(sql)
        wb = Workbook(write_only = True)
        ws = wb.create_sheet()
        ws.append(['日期','教室','应到人数','第一次出勤','第二次出勤','迟到人数','违纪人数','请假人数','早退人数','备注','学院','查早组员','组员姓名','周数'])
        for result in results:
            result = list(result)
            data_temp = [0 for i in range(14)]
            data_temp[0] = result[0] # 日期
            data_temp[1] = result[1]+result[2]+result[3] # 教室
            data_temp[3] = result[4] # json数据
            
            a = data_temp[0] - datetime.timedelta(days=data_temp[0].weekday())
            sql = "SELECT `学院`,`应到人数`,`查早组员`,`姓名`,`周数` FROM `查早排班` WHERE `周起始日期` = '{}' AND `教学楼` = '{}' AND `区号` = '{}' AND `教室编号` = '{}' ORDER BY `教学楼`,`区号`,`教室编号` ASC;".format(a,result[1],result[2],result[3])
            result = self.con_data.execute_query(sql)[0]
            data_temp[2] = result[1] # 应到人数

            unjson = json.loads(data_temp[3])
            data_temp[3] = unjson["第一次出勤"] # 第一次出勤
            data_temp[4] = unjson["第二次出勤"] # 第二次出勤
            data_temp[5] = unjson["迟到人数"] # 迟到人数
            data_temp[6] = unjson["违纪人数"] # 违纪人数
            data_temp[7] = unjson["请假人数"] # 请假人数
            data_temp[8] = unjson["早退人数"] # 早退人数
            data_temp[9] = unjson["备注"] # 备注

            data_temp[10] = result[0] # 学院
            data_temp[11] = result[2] # 查早组员
            data_temp[12] = result[3] # 组员姓名
            data_temp[13] = result[4] # 周数

            ws.append(data_temp.copy())

        wb.save(path)

    def _download_queqin(self, path):# 下载早自习教室缺勤数据
        path = os.path.join(path,"早自习缺勤数据.xlsx")
        print("请输入导出数据的起始日期，例如：2019-05-06")
        while(1):
            try:
                time_start = datetime.datetime.strptime(input(),"%Y-%m-%d")
                break
            except:
                print("日期错误，请重新输入")
        print("已设定起始日期为：{}".format(time_start))
        print("请输入导出数据的结束日期，例如：2019-05-06\n直接回车不输入日期则默认导出一周数据")
        while(1):
            a = input()
            if a=="":
                time_end = time_start+datetime.timedelta(days=6)
                break
            else:
                try:
                    time_end = datetime.datetime.strptime(a,"%Y-%m-%d")
                    break
                except:
                    print("日期错误，请重新输入")
        print("已设定结束日期为：{}".format(time_end))

        sql = "SELECT `日期`,`教学楼`,`区号`,`教室编号`,`缺勤名单`,`提交者` FROM `缺勤人员名单` WHERE `日期` BETWEEN '{}' AND '{}' ORDER BY `教学楼`,`区号`,`教室编号`,`日期` ASC;".format(time_start,time_end)
        results = self.con_data.execute_query(sql)
        wb = Workbook(write_only = True)
        ws = wb.create_sheet()
        ws.append(['日期','教室','姓名','学号','学院','查早组员','组员姓名','周数'])
        for result in results:
            result = list(result)
            data_temp = [0 for i in range(8)]
            data_temp[0] = result[0] # 日期
            data_temp[1] = result[1]+result[2]+result[3] # 教室
            data_temp[2] = result[4] # json数据
            
            a = data_temp[0] - datetime.timedelta(days=data_temp[0].weekday())
            sql = "SELECT `学院`,`应到人数`,`查早组员`,`姓名`,`周数` FROM `查早排班` WHERE `周起始日期` = '{}' AND `教学楼` = '{}' AND `区号` = '{}' AND `教室编号` = '{}' ORDER BY `教学楼`,`区号`,`教室编号` ASC;".format(a,result[1],result[2],result[3])
            result = self.con_data.execute_query(sql)[0]

            data_temp[4] = result[0] # 学院
            data_temp[5] = result[2] # 查早组员
            data_temp[6] = result[3] # 组员姓名
            data_temp[7] = result[4] # 周数

            unjson = json.loads(data_temp[2])
            for xuehao in unjson.keys():
                data_temp[2] = unjson[xuehao] # 姓名
                data_temp[3] = xuehao # 学号
                ws.append(data_temp.copy())

        wb.save(path)

class chake: # 查课类，提供相关数据的上传和下载
    def __init__(self):
        self.con_info = STOS_DB_conn("-p", "./config.json", (True,0))
        self.con_data = STOS_DB_conn("-p", "./config.json", (True,1))

    def __del__(self):
        del self.con_info
        del self.con_data

    def work_info(self, path):
        pass

    def work_schedule(self, path):
        print("请输入该周周一对应的日期，例如：2019-05-06")
        while(1):
            try:
                time_start = datetime.datetime.strptime(input(),"%Y-%m-%d")
                break
            except:
                print("日期错误，请重新输入")
        print("已设定起始日期为：{}".format(time_start))
        time_end = time_start+datetime.timedelta(days=6)
        print("已设定结束日期为：{}".format(time_end))

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(['姓名','学号','编号'])

        exe = pd.read_excel(path)
        count = 1
        for col in range(1,6):
            for row in range(0,3):
                names_group_str = exe.iloc[row,col].split('、')[:-1]
                for name_group in names_group_str:
                    name = name_group.split('(')[0].split(' ')[0]
                    sql = "SELECT `学号` FROM `成员信息` WHERE `姓名` LIKE '%{}%';".format(name)
                    ID = self.con_info.execute_query(sql)[0][0]
                    ws.append([name,ID,count].copy())

                    sql = "UPDATE `查课排班` SET `组员姓名`='{}',`查课组员`='{}' WHERE `日期` BETWEEN '{}' AND '{}' AND `编号` = {};".format(name,ID,time_start,time_end,count)
                    self.con_data.push_query_list(sql)

                    count += 1

        self.con_data.commit_query_list(sql)
        wb.save("_schedule.xlsx")

    def download(self, path):
        path = os.path.join(path,"查课数据.xlsx")
        print("请输入导出数据的起始日期，例如：2019-05-06")
        while(1):
            try:
                time_start = datetime.datetime.strptime(input(),"%Y-%m-%d")
                break
            except:
                print("日期错误，请重新输入")
        print("已设定起始日期为：{}".format(time_start))
        print("请输入导出数据的结束日期，例如：2019-05-06\n直接回车不输入日期则默认导出一周数据")
        while(1):
            a = input()
            if a=="":
                time_end = time_start+datetime.timedelta(days=6)
                break
            else:
                try:
                    time_end = datetime.datetime.strptime(a,"%Y-%m-%d")
                    break
                except:
                    print("日期错误，请重新输入")
        print("已设定结束日期为：{}".format(time_end))

        sql = "SELECT `日期`,`时段与上课周`,`教学楼`,`区号`,`教室编号`,`教室数据`,`提交者`,`编号` FROM `查课数据` WHERE `日期` BETWEEN '{}' AND '{}' ORDER BY `日期`,`编号`,`教学楼`,`区号`,`教室编号` ASC;".format(time_start,time_end)
        results = self.con_data.execute_query(sql)
        wb = Workbook(write_only = True)
        ws = wb.create_sheet()
        ws.append(['日期与时段','课程名称','教室','应到人数','第一次出勤','第二次出勤','第一次违纪','第二次违纪','备注','年级','学院','查课组员','组员姓名'])
        for result in results:
            result = list(result)
            data_temp = [0 for i in range(13)]
            data_temp[0] = result[0]+result[1] # 日期、时段与上课周
            data_temp[2] = result[2]+result[3]+result[4] # 教室
            data_temp[4] = result[5] # json数据
            
            sql = "SELECT `课程名称`,`年级`,`学院`,`应到人数`,`查课组员`,`姓名`, FROM `查课排班` WHERE `编号` = '{}' AND `教学楼` = '{}' AND `区号` = '{}' AND `教室编号` = '{}' ORDER BY `教学楼`,`区号`,`教室编号` ASC;".format(result[7],result[2],result[3],result[4])
            result = self.con_data.execute_query(sql)[0]
            data_temp[1] = result[0] # 课程名称
            data_temp[3] = result[3] # 应到人数

            unjson = json.loads(data_temp[4])
            data_temp[4] = unjson["第一次出勤"] # 第一次出勤
            data_temp[5] = unjson["第二次出勤"] # 第二次出勤
            data_temp[6] = unjson["第一次违纪"] # 第一次违纪
            data_temp[7] = unjson["第二次违纪"] # 第二次违纪
            data_temp[8] = unjson["备注"] # 备注

            data_temp[9] = result[1] # 年级
            data_temp[10] = result[2] # 学院
            data_temp[11] = result[4] # 查课组员
            data_temp[12] = result[5] # 组员姓名

            ws.append(data_temp.copy())

        wb.save(path)