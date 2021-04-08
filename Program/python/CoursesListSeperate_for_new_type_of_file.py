# -*- coding:utf-8 -*-
import datetime
from os.path import join
import re
import os
from numpy.core.fromnumeric import nonzero
from openpyxl.worksheet.worksheet import Worksheet

import pandas as pd
from pandas.core.frame import DataFrame
from pandas.core.indexes.base import Index
import numpy as np

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, named_styles
from openpyxl.styles import colors
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


# 特殊数组需人工更改：院系名称，院系全称，年级
# 请先通过程序提供的选项将无用列删除
# 本程序将处理输入文件并生成新的表单，表单名为"processed"

SchoolName = ['通信', '电子', '材料', '机械', '光电', '自动化', '资源', '计算机', '信息与软件',
              '航空航天', '数学', '物理', '医学', '生命', '经济', '公共', '外国语', '格拉斯哥', '英才']
SchoolName_whole = {
    '通信': '信息与通信工程学院',
    '电子': '电子科学与工程学院',
    '材料': '材料与能源学院',
    '机械': '机械与电气工程学院',
    '光电': '光电科学与工程学院',
    '自动化': '自动化工程学院',
    '资源': '资源与环境学院',
    '计算机': '计算机科学与工程学院',
    '信息与软件': '信息与软件工程学院',
    '航空航天': '航空航天学院',
    '数学': '数学科学学院',
    '物理': '物理学院',
    '医学': '医学院',
    '生命': '生命科学与技术学院',
    '经济': '经济与管理学院',
    '公共': '公共管理学院',
    '外国语': '外国语学院',
    '格拉斯哥': '格拉斯哥学院',
    '英才': '英才实验学院'}
GradeNum = ['2017', '2018', '2019', '2020']


def log_record(log: str, save_dir: str) -> None:  # 用于记录错误信息
    """不返回任何值\n
    接受两个参数：\n
        第一个参数是需要记录的消息字符串\n
        第二个参数是保存的路径，实际存储文件是该路径下的log.log文件"""
    with open(os.path.join(save_dir, 'log.log'), 'a') as f:  # 以“继续添加”模式打开文件，如果不存在则自动创建文件
        f.write(log+'\n\n')  # 两次记录之间有一个空行，方便查看


class OldFileProcessor:
    """全校课表Excel数据读入\n
    提供无用列删除，无用行删除的功能"""
    def __init__(self) -> None:
        self.error = False  # 如果出错则此值为True
        self.fileFullPath = ''  # 需处理文件的路径
        self.fileName = ''  # 需处理文件的名称
        self.sheetName = ''  # 需处理表单的名称
        self.fileDataFrame = pd.DataFrame()  # 读入数据存放处，也是处理后数据的存放处
        self.rowNum = 0  # 数据的行数
        self.colNum = 0  # 数据段列数
        self.usefulColNo = []  # 记录需要保留的列，后续处理数据时将删除无用列

    def read_file(self, fileFullPathAndName: str, sheetName: str) -> None:
        """读取文件，并将内容读入DataFrame"""
        try:
            self.fileFullPath, self.fileName = os.path.split(fileFullPathAndName)
            self.sheetName = sheetName
            # 把Excel数据读入DataFrame
            self.fileDataFrame = pd.DataFrame( pd.read_excel(fileFullPathAndName, sheet_name=self.sheetName) )
            self.rowNum, self.colNum = self.fileDataFrame.shape
            self.error = False
        except Exception as e:
            self.error = True
            print(e)
            log_record('We encounter some errors when open file and read data with Pandas\n\t'+str(e), self.fileFullPath)

    def set_useful_colNo(self, NoList: list) -> None:
        '''设置需要保留的列的序号'''
        self.usefulColNo = NoList

    def show_colName(self, No:list=[]) -> Index:
        '''返回列名'''
        if No==[]:
            return self.fileDataFrame.columns
        else:
            return self.fileDataFrame.columns[No]

    def preprocess_dataFrame(self) -> None:
        '''数据预处理函数\n
        处理内容：\n
            删除无用列\n
            删除空值所在行\n
            根据自定义筛选，去除部分行或列'''
        assert isinstance(self.fileDataFrame, DataFrame)

        # 更新数据帧，保留有用的列
        self.fileDataFrame = self.fileDataFrame.iloc[:, self.usefulColNo]
        # 删除NaN空值，忽略单双周
        self.fileDataFrame = self.fileDataFrame.dropna(how='any')
        # 自定义筛选
        self.CustomFilterFunction()
        # 更新数据索引和行列数
        self.fileDataFrame = self.fileDataFrame.reset_index(drop=True)
        self.rowNum, self.colNum = self.fileDataFrame.shape
    
    def CustomFilterFunction(self) -> None: # 筛选校区
        '''自定义函数：\n
            筛选校区，保留清水河校区课程\n
            去除上课人数少于10人的教室'''
        # 筛选出“校区”列内容为“清水河校区”的行
        self.fileDataFrame = self.fileDataFrame[self.fileDataFrame['校区'].isin(['清水河校区'])]
        self.fileDataFrame = self.fileDataFrame[~self.fileDataFrame['上课人数'].isin(range(10))]

class NewFileProcessor:
    '''处理预处理后的文件并存储\n
    提供对上课地点、时间、年级等的筛除功能'''
    def __init__(self) -> None:
        self.error = False
        self.fileFullPath = ''
        self.fileName = ''
        self.filebook = openpyxl.Workbook()
        # self.filesheet # this prop create in specific function
        self.sheetName = 'processed'
        self.insertDataFrame = pd.DataFrame()
        self.rowNum = 0 # 数据的行数
        self.colNum = 0 # 数据段列数

    def create_from_OldFileProcessor(self, old_file_processor : OldFileProcessor) -> None:
        '''的'''
        try:
            self.fileFullPath = old_file_processor.fileFullPath
            self.fileName = old_file_processor.fileName
            self.insertDataFrame = old_file_processor.fileDataFrame
            self.rowNum = old_file_processor.rowNum # 数据的行数
            self.colNum = old_file_processor.colNum # 数据段列数
            self.error = old_file_processor.error
            if self.error: raise Exception
            self.filebook = openpyxl.load_workbook( os.path.join(self.fileFullPath,self.fileName) )
            self.filesheet = self.filebook.create_sheet(self.sheetName)
            self.error = False
        except Exception as e:
            self.error = True
            print(e)
            log_record('We encounter some errors when import data from original data and create new sheet in file\n\t'+str(e), self.fileFullPath)

    def process_data(self) -> None:
        '''正式数据处理'''
        assert isinstance(self.insertDataFrame,DataFrame
        )
        # 读入待处理数据
        dataProcessor = dataNormalization(self.insertDataFrame)
        # 处理数据后返回结果
        self.insertDataFrame = dataProcessor()

    def save_data_and_file(self) -> None:
        '''将数据添加至Excel并保存'''
        assert isinstance(self.filesheet, Worksheet)

        try:
            # 按行添加数据至Excel
            for one_row in dataframe_to_rows(self.insertDataFrame,index=False):
                self.filesheet.append(one_row)
            
            # 按行索引
            for row_index in range(1,self.rowNum+2):
                # 设定行高
                self.filesheet.row_dimensions[row_index].height = 18
                # 在行索引基础上索引列，获取每一个单元格
                for cell_index in range(1,self.colNum+1):
                    # 设置单元格对齐方式，并且对第一行的单元格设置底色
                    if row_index==1:
                        self.filesheet.cell(row=row_index,column=cell_index).fill = PatternFill("solid", fgColor="AAAAAA")
                    self.filesheet.cell(row=row_index,column=cell_index).alignment = Alignment(horizontal='center',vertical='center')
            
            # 按列索引
            for col_index in range(1,self.colNum+1):
                col_str = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']
                col_width = [6,6,25,50,13,9,30,7,12,35]
                # 设置列宽
                self.filesheet.column_dimensions[col_str[col_index-1]].width = col_width[col_index-1]
            
            # 存储文件
            self.filebook.save( os.path.join(self.fileFullPath,self.fileName) )
        except Exception as e:
            self.error = True
            print(e)
            log_record('We encounter some errors when save data into sheet and save file\n\t'+str(e), self.fileFullPath)

class dataNormalization:
    '''正式处理数据的类'''
    def __init__(self, dataFrame : DataFrame) -> None:
        self.oriDataFrame = dataFrame # 保留一份原始数据
        self.rowNum, self.colNum = dataFrame.shape
        self.tmpDataFrame = DataFrame() # 期间可能用到的临时数据存放点
        self.finDataFrame = DataFrame() # 最终处理完返回的数据
        self.error = False

    def __call__(self) -> DataFrame:
        self.normalization()
        return self.finDataFrame

    def normalization(self):
        '''数据处理程序'''
        self.finDataFrame = self.oriDataFrame.copy()
        for i in self.finDataFrame.index:
            one_row = self.oriDataFrame.loc[i]
            cr,gd,sc = self.classroom_processor(one_row['上课教室']),self.grade_processor(one_row['年级']),self.school_processor(one_row['上课院系'])
            if False in [cr,gd,sc]:
                self.finDataFrame = self.finDataFrame.drop(index=i)
            else:
                self.finDataFrame.loc[i,'上课教室'] = cr
                self.finDataFrame.loc[i,'年级'] = gd
                self.finDataFrame.loc[i,'上课院系'] = sc
        self.finDataFrame = self.finDataFrame.reset_index(drop=True)

    def classroom_processor(self,cr):
        '''对上课教室列的处理'''
        pattern = re.compile(r'立人楼[A-Za-z]-?[0-9]{3}|品学楼[A-Za-z]-?[0-9]{3}-?[A-Za-z]?')
        location_set = set(pattern.findall(cr)) # 获取符合要求的所有教室，并做去重复处理
        if len(location_set)==1: # 如果只有一个上课地点则保留所在行，理由：每一行对应一个老师周内的一个时间，只可能有一个上课地点，多个地点说明会根据周的不同改变，但所提供数据无法提供具体从哪一周起改变教室，故做删除处理
            return location_set.pop()
        else:
            return False

    def grade_processor(self,gd):
        '''对年级列的处理'''
        count = 0
        name = ''
        for n in GradeNum: # 如果只有一个年级则保留，多年级则删去行
            if gd.find(n) != -1:
                count += 1
                name = n
        if count==1:
            return name
        else:
            return False

    def school_processor(self,sc):
        '''对学院列的处理'''
        count = 0
        name = ''
        for n in SchoolName: # 只有一个学院则保留，多学院则删除行
            if sc.find(n) != -1:
                count += 1
                name = SchoolName_whole[n]
        if count==1:
            return name
        else:
            return False


def main():
    pathandname = input('输入文件绝对路径和文件名：\n')
    sheetname = input('输入需要处理的表单名称：\n')
    OLD = OldFileProcessor()
    OLD.read_file(pathandname,sheetname)

    print("以下列出了Excel的列名称，请选择需要保留的列编号，编号用空格隔开")
    print(list(enumerate(OLD.show_colName())))
    No = list(map(lambda x:int(x), input("输入编号：\n").split(' ')))
    OLD.set_useful_colNo(No)
    No = list(zip(No,OLD.show_colName(No)))
    print("你选择的编号与对应的列名称是：\n",No)

    print("\n准备：去除无用的列和值为空的行")
    OLD.preprocess_dataFrame()
    print("\t完成：去除无用的列和值为空的行")

    print("\n预处理已完成")
    print("\n准备：数据转储到正式处理程序")
    NEW = NewFileProcessor()
    NEW.create_from_OldFileProcessor(OLD)
    print("\t完成：数据转储到正式处理程序")
    
    print("\n准备：正式处理数据")
    NEW.process_data()
    print("\t完成：正式处理数据")
    a = NEW.insertDataFrame

    NEW.save_data_and_file()
    print("\n数据已保存至processed表单")


if __name__ == "__main__":
    startTime = datetime.datetime.now()
    
    main()

    endTime = datetime.datetime.now()
    print( "Process finished in {}".format(endTime-startTime) )
