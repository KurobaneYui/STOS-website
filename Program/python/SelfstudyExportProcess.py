# -*- coding:utf-8 -*-
import datetime
import time
import sys
import os
import json
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from Frame.python3.BaseComponents.DatabaseConnector import DatabaseConnector
from openpyxl.utils import get_column_letter

# 写入Excel文件数据


def writedata(path: str, database: DatabaseConnector,
              startDate: datetime.datetime, endDate: datetime.datetime):
    '''
    path: 文件保存路径
    database: MySQL连接
    selfstudy_start_data 起始日期
    selfstudy_end_data 结束日期
    '''
    # 创建excel，建立第一个表单“数据”和第二个表单“名单”
    wb = openpyxl.Workbook()
    ws_sheet1 = wb.active
    ws_sheet1.title = '数据'
    ws_sheet2 = wb.create_sheet('缺勤')
    assert isinstance(ws_sheet2, Worksheet)
    ws_sheet3 = wb.create_sheet('请假')
    assert isinstance(ws_sheet3, Worksheet)

    # 从排班中获取给定日期范围内的排班，并获取教室、学院、编号、检查人员姓名学号、所属组等信息
    # 便利每一个排班，利用编号搜索数据
    # # 利用编号查找早自习数据和人员名单数据，只取第一个可行解
    # # 将获取到的数据用json解包，并更新入排班的字典内
    # 将每一个排班字典按规范导入excel
    database.execute("SELECT selfstudy_id, campus, school_name, student_supposed, classroom_name, actual_student_id, actual_student_name, actual_student_department_name, date \
                    FROM SelfstudyCheckActualView \
                    WHERE date>=%s AND date<=%s;",
                     data=(startDate, endDate))
    schedules = database.fetchall()
    for one_schedule in schedules:
        DBAffectedRows = database.execute("SELECT check_result, groupleader_recheck AS recheck, remark \
                        FROM SelfstudyCheckData  \
                        WHERE selfstudy_id=%s \
                        ORDER BY submission_time DESC \
                        LIMIT 1;",
                                          data=(one_schedule["selfstudy_id"],))
        record = database.fetchall()
        if DBAffectedRows != 0:
            record = record[0]
            one_schedule.update({
                "recheck": record["recheck"] == 1,
                "recheck_remark": record["remark"]
            })
            one_schedule.update(json.loads(record["check_result"]))
        else:
            one_schedule.update({
                "recheck": None,
                "recheck_remark": None,
                "student_supposed": None,
                "firstPresent": None,
                "absent": None,
                "secondPresent": None,
                "leaveEarly": None,
                "askForLeave": None,
                "absent": None,
                "remark": None
            })

        DBAffectedRows = database.execute("SELECT check_result \
                        FROM SelfstudyCheckAbsent \
                        WHERE selfstudy_id = %s \
                        ORDER BY submission_time DESC \
                        LIMIT 1;",
                                          data=(one_schedule["selfstudy_id"],))
        absentList = database.fetchall()
        if DBAffectedRows != 0:
            absentList = absentList[0]
            one_schedule.update({
                "absentList": json.loads(absentList["check_result"])
            })
        else:
            one_schedule.update({
                "absentList": []
            })

        DBAffectedRows = database.execute("SELECT check_result \
                        FROM SelfstudyCheckAskForLeave \
                        WHERE selfstudy_id = %s \
                        ORDER BY submission_time DESC \
                        LIMIT 1;",
                                          data=(one_schedule["selfstudy_id"],))
        askForLeaveList = database.fetchall()
        if DBAffectedRows != 0:
            askForLeaveList = askForLeaveList[0]
            one_schedule.update({
                "askForLeaveList": json.loads(askForLeaveList["check_result"])
            })
        else:
            one_schedule.update({
                "askForLeaveList": []
            })

    ws_sheet1.append(['日期', '教室', '校区', '学院', '应到人数', '第一次出勤', '迟到人数',
                     '第二次出勤', '早退人数', '请假人数', '备注', '组长确认', '组长备注', '查早组员', '组员学号', '所属组'])
    ws_sheet2.append(['日期', '教室', '校区', '学院', '缺勤人姓名',
                     '缺勤人学号', '查早组员', '组员学号', '所属组'])
    ws_sheet3.append(['日期', '教室', '校区', '学院', '请假人姓名',
                     '请假人学号', '事由', '查早组员', '组员学号', '所属组'])

    absentOffset = 2
    askForLeaveOffset = 2
    for order, one_schedule in enumerate(schedules, start=2):
        ws_sheet1.cell(row=order, column=1, value=one_schedule['date'])
        ws_sheet1.cell(row=order, column=2,
                       value=one_schedule['classroom_name'])
        ws_sheet1.cell(row=order, column=3, value=one_schedule['campus'])
        ws_sheet1.cell(row=order, column=4, value=one_schedule['school_name'])
        ws_sheet1.cell(row=order, column=5,
                       value=one_schedule['student_supposed'])
        ws_sheet1.cell(row=order, column=6, value=one_schedule['firstPresent'])
        ws_sheet1.cell(row=order, column=7, value=one_schedule['absent'])
        ws_sheet1.cell(row=order, column=8,
                       value=one_schedule['secondPresent'])
        ws_sheet1.cell(row=order, column=9, value=one_schedule['leaveEarly'])
        ws_sheet1.cell(row=order, column=10, value=one_schedule['askForLeave'])
        ws_sheet1.cell(row=order, column=11, value=one_schedule['remark'])
        ws_sheet1.cell(row=order, column=12,
                       value='是' if one_schedule['recheck'] in [1, '1'] else '否')
        ws_sheet1.cell(row=order, column=13,
                       value=one_schedule['recheck_remark'])
        ws_sheet1.cell(row=order, column=14,
                       value=one_schedule['actual_student_name'])
        ws_sheet1.cell(row=order, column=15,
                       value=one_schedule['actual_student_id'])
        ws_sheet1.cell(row=order, column=16,
                       value=one_schedule['actual_student_department_name'])

        for one_absent_student in one_schedule["absentList"]:
            ws_sheet2.cell(row=absentOffset, column=1,
                           value=one_schedule['date'])
            ws_sheet2.cell(row=absentOffset, column=2,
                           value=one_schedule['classroom_name'])
            ws_sheet2.cell(row=absentOffset, column=3,
                           value=one_schedule['campus'])
            ws_sheet2.cell(row=absentOffset, column=4,
                           value=one_schedule['school_name'])

            ws_sheet2.cell(row=absentOffset, column=5,
                           value=one_absent_student['student_name'])
            ws_sheet2.cell(row=absentOffset, column=6,
                           value=one_absent_student['student_id'])

            ws_sheet2.cell(row=absentOffset, column=7,
                           value=one_schedule['actual_student_name'])
            ws_sheet2.cell(row=absentOffset, column=8,
                           value=one_schedule['actual_student_id'])
            ws_sheet2.cell(row=absentOffset, column=9,
                           value=one_schedule['actual_student_department_name'])

            absentOffset += 1

        for one_askForLeave_student in one_schedule["askForLeaveList"]:
            ws_sheet3.cell(row=askForLeaveOffset, column=1,
                           value=one_schedule['date'])
            ws_sheet3.cell(row=askForLeaveOffset, column=2,
                           value=one_schedule['classroom_name'])
            ws_sheet3.cell(row=askForLeaveOffset, column=3,
                           value=one_schedule['campus'])
            ws_sheet3.cell(row=askForLeaveOffset, column=4,
                           value=one_schedule['school_name'])

            ws_sheet3.cell(row=askForLeaveOffset, column=5,
                           value=one_askForLeave_student['student_name'])
            ws_sheet3.cell(row=askForLeaveOffset, column=6,
                           value=one_askForLeave_student['student_id'])
            ws_sheet3.cell(row=askForLeaveOffset, column=7,
                           value=one_askForLeave_student['reason'])

            ws_sheet3.cell(row=askForLeaveOffset, column=8,
                           value=one_schedule['actual_student_name'])
            ws_sheet3.cell(row=askForLeaveOffset, column=9,
                           value=one_schedule['actual_student_id'])
            ws_sheet3.cell(row=askForLeaveOffset, column=10,
                           value=one_schedule['actual_student_department_name'])

            askForLeaveOffset += 1

    # 保存文件
    wb.save(path)


if __name__ == '__main__':
    '''
    以下是所需参数及其顺序和类型:
        path:str 文件保存路径目录
        startDate:datetime 起始日期
        endDate:datetime 结束日期
    '''
    try:
        input_len = len(sys.argv)-1
        if input_len == 3:
            session = DatabaseConnector()
            session.startCursor()

            path = os.path.join(sys.argv[1], str(time.time()))
            os.makedirs(path, exist_ok=True)
            path = os.path.join(path, '早自习数据.xlsx')
            params = {
                'path': path,
                'database': session,
                'startDate': datetime.datetime.strptime(sys.argv[2], '%Y-%m-%d'),
                'endDate': datetime.datetime.strptime(sys.argv[3], '%Y-%m-%d')
            }
            writedata(**params)
        else:
            raise Exception('程序输入参数数量与需求不一致，获得了{}个参数'.format(input_len))
    except Exception as e:
        errors = {'ReturnCode': '417', 'ReturnString': '程序出错',
                  'ShowMessage': repr(e), 'Data': ''}
        print(json.dumps(errors, ensure_ascii=False))
    else:
        datetime.datetime.strptime('2013-2-3', '%Y-%m-%d')
        print(json.dumps({'ReturnCode': '200', 'ReturnString': '成功',
              'ShowMessage': '', 'Data': path}, ensure_ascii=False))
