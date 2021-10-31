# -*- coding:utf-8 -*-

import re
import os
import sys
import json
import datetime
from typing import Tuple
from mysql import connector
from mysql.connector.connection import MySQLConnection
from mysql.connector.cursor import CursorBase

import pandas as pd
from pandas import DataFrame
from pandas import Series
from pandas import Index
import numpy as np
from numpy import isin, nonzero

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, named_styles
from openpyxl.styles import colors
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


# 特殊数组需人工更改：院系名称，院系全称，年级
# 请先通过程序提供的选项将无用列删除
# 本程序将处理输入文件并生成新的表单，表单名为"processed"

SchoolName = ['通信', '电子', '材料', '机械', '光电', '自动化', '资源', '计算机', '信息与软件',
              '航空航天', '数学', '物理', '医学', '生命', '经济', '公共', '外国语', '格拉斯哥', '英才']
SchoolName_whole = {
    '通信': '信息与通信工程学院',
    '电子': '电子科学与工程学院',
    '材料': '材料与能源学院',
    '机械': '机械与电气工程学院',
    '光电': '光电科学与工程学院',
    '自动化': '自动化工程学院',
    '资源': '资源与环境学院',
    '计算机': '计算机科学与工程学院',
    '信息与软件': '信息与软件工程学院',
    '航空航天': '航空航天学院',
    '数学': '数学科学学院',
    '物理': '物理学院',
    '医学': '医学院',
    '生命': '生命科学与技术学院',
    '经济': '经济与管理学院',
    '公共': '公共管理学院',
    '外国语': '外国语学院',
    '格拉斯哥': '格拉斯哥学院',
    '英才': '英才实验学院'}
GradeNum = ['2017', '2018', '2019', '2020']


def log_record(log: str, save_dir: str) -> None:  # 用于记录错误信息
    """不返回任何值\n
    接受两个参数：\n
        第一个参数是需要记录的消息字符串\n
        第二个参数是保存的路径，实际存储文件是该路径下的log.log文件"""
    with open(os.path.join(save_dir, 'log.log'), 'a+') as f:  # 以“继续添加”模式打开文件，如果不存在则自动创建文件
        f.write(log+'\n\n')  # 两次记录之间有一个空行，方便查看


class OldFileProcessor:
    """全校课表Excel数据读入\n
    提供无用列删除，无用行删除的功能，提供简单的自定义筛除功能，如筛选校区和上课人数大于10人的课等"""
    def __init__(self) -> None:
        self.error = False  # 如果出错则此值为True
        self.fileFullPath = ''  # 需处理文件的路径
        self.fileName = ''  # 需处理文件的名称
        self.sheetName = ''  # 需处理表单的名称
        self.fileDataFrame = pd.DataFrame()  # 读入数据存放处，也是处理后数据的存放处
        self.rowNum = 0  # 数据的行数
        self.colNum = 0  # 数据段列数
        self.usefulColNo = []  # 记录需要保留的列，后续处理数据时将删除无用列

    def read_file(self, fileFullPathAndName: str, sheetName: str) -> None:
        """读取文件，并将内容读入DataFrame"""
        try:
            self.fileFullPath, self.fileName = os.path.split(fileFullPathAndName)
            self.sheetName = sheetName
            # 把Excel数据读入DataFrame
            self.fileDataFrame = pd.DataFrame( pd.read_excel(fileFullPathAndName, sheet_name=self.sheetName) )
            self.rowNum, self.colNum = self.fileDataFrame.shape
            self.error = False
        except Exception as e:
            self.error = True
            print(e)
            log_record('We encounter some errors when open file and read data with Pandas\n\t'+str(e), self.fileFullPath)

    def set_useful_colNo(self, NoList: list) -> None:
        """设置需要保留的列的序号"""
        self.usefulColNo = NoList

    def show_colName(self, No:list=[]) -> Index:
        """返回列名"""
        assert isinstance(self.fileDataFrame, DataFrame)
        if No==[]:
            return self.fileDataFrame.columns
        else:
            return self.fileDataFrame.iloc[:,No].columns

    def preprocess_dataFrame(self) -> None:
        """数据预处理函数\n
        处理内容：\n
            删除无用列\n
            删除空值所在行\n
            根据自定义筛选，去除部分行或列"""
        assert isinstance(self.fileDataFrame, DataFrame)

        # 更新数据帧，保留有用的列
        self.fileDataFrame = self.fileDataFrame.iloc[:, self.usefulColNo]
        # 删除NaN空值，忽略单双周
        self.fileDataFrame = self.fileDataFrame.dropna(how='any')
        # 自定义筛选
        self.CustomFilterFunction()
        # 更新数据索引和行列数
        self.fileDataFrame = self.fileDataFrame.reset_index(drop=True)
        self.rowNum, self.colNum = self.fileDataFrame.shape
    
    def CustomFilterFunction(self) -> None: # 筛选校区和上课人数过少的课
        """自定义函数：\n
            筛选校区，保留清水河校区课程\n
            去除上课人数少于10人的教室"""
        # 筛选出“校区”列内容为“清水河校区”的行
        self.fileDataFrame = self.fileDataFrame[self.fileDataFrame['校区'].isin(['清水河校区'])]
        self.fileDataFrame = self.fileDataFrame[~self.fileDataFrame['上课人数'].isin(range(10))]


class NewFileProcessor:
    """处理预处理后的文件并存储\n
    提供对上课地点、时间、年级等的筛除功能，功能定义在单独的类(dataNormalization)中"""
    def __init__(self) -> None:
        self.error = False
        self.fileFullPath = ''
        self.fileName = ''
        self.filebook = openpyxl.Workbook()
        # self.filesheet # this prop create in specific function
        self.sheetName = 'processed'
        self.insertDataFrame = pd.DataFrame()
        self.rowNum = 0 # 数据的行数
        self.colNum = 0 # 数据段列数

    def create_from_OldFileProcessor(self, old_file_processor : OldFileProcessor) -> None:
        """从预处理程序读取数据"""
        try:
            assert isinstance(old_file_processor,OldFileProcessor)
            self.fileFullPath = old_file_processor.fileFullPath
            self.fileName = old_file_processor.fileName
            self.insertDataFrame = old_file_processor.fileDataFrame
            self.rowNum = old_file_processor.rowNum # 数据的行数
            self.colNum = old_file_processor.colNum # 数据段列数
            self.error = old_file_processor.error
            if self.error: raise Exception
            self.filebook = openpyxl.load_workbook( os.path.join(self.fileFullPath,self.fileName) )
            self.filesheet = self.filebook.create_sheet(self.sheetName)
            self.error = False
        except Exception as e:
            self.error = True
            print(e)
            log_record('We encounter some errors when import data from original data and create new sheet in file\n\t'+str(e), self.fileFullPath)

    def process_data(self) -> None:
        """正式数据处理"""
        assert isinstance(self.insertDataFrame,DataFrame)
        # 读入待处理数据，并返回处理结果
        self.insertDataFrame = dataNormalization(self.insertDataFrame)()

    def save_data_and_file(self) -> None:
        """将数据添加至Excel并保存"""
        assert isinstance(self.filesheet, Worksheet)

        try:
            # 按行添加数据至Excel
            for one_row in dataframe_to_rows(self.insertDataFrame,index=False):
                self.filesheet.append(one_row)
            
            # 按行索引
            for row_index in range(1,self.rowNum+2):
                # 设定行高
                self.filesheet.row_dimensions[row_index].height = 18
                # 在行索引基础上索引列，获取每一个单元格
                for cell_index in range(1,self.colNum+1):
                    # 设置单元格对齐方式，并且对第一行的单元格设置底色
                    if row_index==1:
                        self.filesheet.cell(row=row_index,column=cell_index).fill = PatternFill("solid", fgColor="AAAAAA")
                    self.filesheet.cell(row=row_index,column=cell_index).alignment = Alignment(horizontal='center',vertical='center')
            
            # 按列索引
            for col_index in range(1,self.colNum+1):
                col_str = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']
                col_width = [6,6,25,50,13,9,30,7,12,35]
                # 设置列宽
                self.filesheet.column_dimensions[col_str[col_index-1]].width = col_width[col_index-1]
            
            # 存储文件
            self.filebook.save( os.path.join(self.fileFullPath,self.fileName) )
        except Exception as e:
            self.error = True
            print(e)
            log_record('We encounter some errors when save data into sheet and save file\n\t'+str(e), self.fileFullPath)


class dataNormalization:
    """正式处理数据的类\n
    提供对数据的筛选、去重、去除不符合要求的课程等高级筛选功能"""
    def __init__(self, dataFrame : DataFrame) -> None:
        self.oriDataFrame = dataFrame # 保留一份原始数据
        self.rowNum, self.colNum = dataFrame.shape
        self.tmpDataFrame = DataFrame() # 期间可能用到的临时数据存放点
        self.finDataFrame = DataFrame() # 最终处理完返回的数据
        self.error = False

    def __call__(self) -> DataFrame:
        self.normalization()
        return self.finDataFrame

    def normalization(self):
        """数据处理程序"""
        self.finDataFrame = self.oriDataFrame.copy()
        for i in self.finDataFrame.index:
            one_row = self.oriDataFrame.loc[i]
            cr,gd,sc = self.classroom_processor(one_row['上课教室']),self.grade_processor(one_row['年级']),self.school_processor(one_row['上课院系'])
            if False in [cr,gd,sc]:
                self.finDataFrame = self.finDataFrame.drop(index=i)
            else:
                self.finDataFrame.at[i,'上课教室'] = cr
                self.finDataFrame.at[i,'年级'] = gd
                self.finDataFrame.at[i,'上课院系'] = sc
        self.finDataFrame = self.finDataFrame.reset_index(drop=True)

    def classroom_processor(self,cr):
        """对上课教室列的处理"""
        pattern = re.compile(r'立人楼[A-Za-z]-?[0-9]{3}|品学楼[A-Za-z]-?[0-9]{3}-?[A-Za-z]?')
        location_set = set(pattern.findall(cr)) # 获取符合要求的所有教室，并做去重复处理
        if len(location_set)==1: # 如果只有一个上课地点则保留所在行，理由：每一行对应一个老师周内的一个时间，只可能有一个上课地点，多个地点说明会根据周的不同改变，但所提供数据无法提供具体从哪一周起改变教室，故做删除处理
            return location_set.pop()
        else:
            return False

    def grade_processor(self,gd):
        """对年级列的处理"""
        count = 0
        name = ''
        for n in GradeNum: # 如果只有一个年级则保留，多年级则删去行
            if gd.find(n) != -1:
                count += 1
                name = n
        if count==1:
            return name
        else:
            return False

    def school_processor(self,sc):
        """对学院列的处理"""
        count = 0
        name = ''
        for n in SchoolName: # 只有一个学院则保留，多学院则删除行
            if sc.find(n) != -1:
                count += 1
                name = SchoolName_whole[n]
        if count==1:
            return name
        else:
            return False


class UpLoader:
    """
    数据库导入程序，将DataFrame格式的课程信息导入数据库
    """
    def __init__(self, oriData: NewFileProcessor, session: MySQLConnection, date: str) -> None:
        """
        读入数据：从正式处理程序里读取DataFrame格式的数据
        """
        assert isinstance(oriData, NewFileProcessor)
        assert isinstance(oriData.insertDataFrame, DataFrame)
        self.oriData = oriData.insertDataFrame.copy() # 要上传的数据（没有转化格式）
        self.oriData_rows, self.oriData_cols = self.oriData.shape # 获取数据尺寸
        self.path = oriData.fileFullPath
        self.session = session # 数据库连接
        self.cur = session.cursor() # 数据库游标
        self.startDate = date # 第一周对应的日期

    def process_upload(self):
        """
        将处理后的数据上传至数据库
        """
        self.cur.execute("select 课程编号  from `全校查课信息` order by 课程编号 DESC limit 1;")
        course_code = self.cur.fetchone() + 1
        for row in self.oriData.index:
            S = self.oriData.loc[row]
            assert isinstance(S, Series)
            classroom_list, course_list = self.pre_process(S, course_code)
            if classroom_list != []:
                sql = "INSERT INTO `教室信息`(`教室编号`, `校区`, `教学楼`, `区域`, `牌号`,`容纳人数`)VALUES(%s,%s,%s,%s,%s,0);"
                try:
                    self.cur.executemany(sql,classroom_list)
                except Exception as e:
                    log_record("插入教室信息时出错：{}，教室如下：{}".format(e, classroom_list), self.path)
            if course_list != []:
                sql = "INSERT INTO `全校查课信息`(`课程编号`, `学院编号`, `年级`, `课程名称`, `教室编号`,`日期`,`时段`, `应到人数`)\
                        VALUES(%s,%s,%s,%s,%s,%s,%s,%s);"
                try:
                    self.cur.executemany(sql,course_list)
                except Exception as e:
                    log_record("插入课程信息时出错：{}，课程如下：{}".format(e, classroom_list), self.path)
                else:
                    course_code = course_list[-1][0] + 1
        self.session.commit()
    
    def search_schoolCode_classroomCode(self, schoolName: str, classroomName: str) -> tuple:
        """
        搜索学院编号和教室编号
        """
        self.cur.execute("select 学院编号 from 学院信息 where 学院名称='{}';".format(schoolName))
        school_code = self.cur.fetchone()[0]
        
        self.cur.execute("select 教室编号 from 教室信息 where 教学楼='{}' and 区域='{}' and 牌号='{}';".format(self.classroom_split(classroomName)))
        classroom_code = self.cur.fetchone()[0]
        
        return school_code, classroom_code

    def pre_process(self, one_record: Series, course_code: int) -> tuple:
        """
        将数据拆分合并为上传数据库需要的形式，如果教室数据已存在，则不添加这个教室
        """
        processedData = {'添加教室': [], '添加课程': []}

        school_code, classroom_code = self.search_schoolCode_classroomCode(one_record['上课院系'], one_record['上课教室'])
        assert school_code!=False

        grade = one_record['年级']
        course_name = one_record['课程名称']
        dates = self.date_transform(one_record['起止周'], one_record['星期'])
        interval = one_record['时段']
        num_expect = int(one_record['上课人数'])
        
        if classroom_code==False:
            classroom_code = self.transform_classname_to_code(one_record['上课教室'])
            classroom_campus = one_record['校区']
            classroom_building, classroom_area, classroom_addr = self.classroom_split(one_record['上课教室'])

            processedData['添加教室'].append((classroom_code, classroom_campus, classroom_building, classroom_area, classroom_addr))
        processedData['添加课程'] = [(course_code+i, school_code, grade, course_name, classroom_code, date, interval, num_expect) for i, date in enumerate(dates)]
        
        return processedData['添加教室'], processedData['添加课程']

    def classroom_split(self, classroom: str) -> tuple:
        """
        将完整的教室字符串拆分为独立的部分，分别为：教学楼、区域、牌号
        """
        classroom1 = re.compile(r'立人楼|品学楼|第二教学楼').findall(classroom)[0] #FIXME:修改匹配项以适配沙河校区
        classroom_ = ''.join(re.compile(r'[a-zA-Z]?-?\d{3}-?[a-zA-Z]?').findall(classroom)[0].split('-')) #FIXME:修改匹配项以适配沙河校区
        if ('a'<=classroom_[0]<='z') or ('A'<=classroom_[0]<='Z'):
            classroom2 = classroom[0]
            classroom3 = classroom_[1:]
        elif ('0'<=classroom_[0]<='9'):
            classroom2 = '-'
            classroom3 = classroom_
        else:
            raise Exception
        return classroom1,classroom2,classroom3

    def transform_classname_to_code(self, class_name: str) -> int:
        """
        ### 根据完整的教室字符串处理得到对应的教室编号

        第一位：1清水河；2沙河

        第二位：1品学楼；2立人楼 or 1第一教学楼；2第二教学楼

        第三位：0-；1A；2B；3C

        第四~六位：教室编号的数字部分

        第七位：教室编号的分片（字母）部分。0无；1A；2B；3C
        """
        building, area, code = self.classroom_split(class_name)
        
        class_code = ''
        
        if building == '品学楼':
            class_code += '11'
        elif building == '立人楼':
            class_code += '12'
        elif building == '第二教学楼':
            class_code += '22'
        
        if area == '-':
            class_code += '0'
        elif area == 'A' or area == 'a':
            class_code += '1'
        elif area == 'B' or area == 'b':
            class_code += '2'
        elif area == 'C' or area == 'c':
            class_code += '3'
        
        if len(code) == 3:
            class_code += str(code) + '0'
        else:
            class_code += str(code)[:3]
            code_ = str(code)[3]
            if code_ == 'A' or code_ == 'a':
                class_code += '1'
            elif code_ == 'B' or code_ == 'b':
                class_code += '2'
            elif code_ == 'C' or code_ == 'c':
                class_code += '3'

        return int(class_code)
    
    def date_transform(self, week: str, weekday: str) -> list:
        """
        转换起止周和星期，变为具体的日期（datetime）
        """
        days = []
        first_date = datetime.datetime.strptime(self.startDate, '%Y-%m-%d')

        pattern = re.compile(r"\d+-?\d*[单双]?")
        weeks = (pattern.findall(week))
        for week_group in weeks:
            if week_group[-1] == '单':
                odd_even = 1
                week_group = week_group[:-1]
            elif week_group[-1] == '双':
                odd_even = 2
                week_group = week_group[:-1]
            else:
                odd_even = 0
            if '-' in week_group:
                start_week, end_week = week_group.split('-')
            else:
                start_week = end_week = week_group
            for i in range(start_week, end_week+1):
                if odd_even == 1 and i % 2 == 0: continue
                if odd_even == 2 and i % 2 == 1: continue
                delta = datetime.timedelta(weeks=i-1, days=int(weekday)-1)
                days.append(first_date + delta)
        return days


if __name__ == "__main__":
    """
    一下是所需参数及其顺序和类型:
        conf:str 配置文件路径
        path:str 待处理和导入的Excel文件路径
        sheetName:str 待处理数据所在表单的名称
        mode:int 模式：0返回表头，1存入数据库
        date:str 如果选择模式1，则需要提供这个参数。教学列表里“第一周”对应的日期
        ...:int(s) 如果选择模式1，则补充需要保留的表头编号，每个编号是一个参数，均要求是整数
    """
    try:
        input_len = len(sys.argv)-1
        if input_len>=4:
            with open(sys.argv[1], 'r') as f:
                with connector.connect(**json.load(f)) as session:
                    assert isinstance(session, MySQLConnection)
                    path = sys.argv[2]
                    mode = int(sys.argv[3])
                    sheetName = sys.argv[4]
                    date_useful_cols = [sys.argv[i] for i in range(5,input_len+1)]
                    if mode==0:
                        OLD = OldFileProcessor() # 初始化预处理
                        OLD.read_file(path,sheetName) # 设置文件和表单
                        returns = json.dumps(dict(enumerate(OLD.show_colName()))) # 显示列名
                        print(json.dumps({'ReturnCode': '200', 'ReturnString': '成功', 'ShowMessage': '', 'Data': returns}, ensure_ascii=False))
                    elif mode==1:
                        date = date_useful_cols[0]
                        useful_cols = [int(i) for i in date_useful_cols[1:]]

                        OLD = OldFileProcessor() # 初始化预处理
                        OLD.read_file(path,sheetName) # 设置文件和表单
                        OLD.set_useful_colNo(useful_cols) # 设置需要保留的列
                        OLD.preprocess_dataFrame() # 预处理
                        log_record("预处理完成", os.path.split(path)[0])
                        NEW = NewFileProcessor() # 初始化正式处理
                        NEW.create_from_OldFileProcessor(OLD) # 从预处理结果读取数据
                        NEW.process_data() # 正式处理
                        log_record("正式处理完成", os.path.split(path)[0])
                        uploader = UpLoader(NEW, session, date) # 初始化上传程序并从正式处理结果读取数据
                        del OLD, NEW
                        uploader.process_upload() # 整理数据为合适的格式并上传
                        log_record("上传完成", os.path.split(path)[0])
                        print(json.dumps({'ReturnCode': '200', 'ReturnString': '成功', 'ShowMessage': '', 'Data': ''}, ensure_ascii=False))
                    else:
                        raise Exception('模式选择错误，选择的模式是{}，而该模式不存在'.format(mode))
        else:
            raise Exception('程序输入参数数量与需求不一致，获得了{}个参数'.format(input_len))
    except Exception as e:
        import json
        errors = {'ReturnCode': '417', 'ReturnString': '程序出错', 'ShowMessage': e, 'Data': ''}
        print(json.dumps(errors, ensure_ascii=False))
    else:
        print(json.dumps({'ReturnCode': '200', 'ReturnString': '成功', 'ShowMessage': '', 'Data': ''}, ensure_ascii=False))