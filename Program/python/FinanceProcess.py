# -*- coding:utf-8 -*-
import datetime
import time
import sys
import os
import json
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from Frame.python3.DatabaseConnector import DatabaseConnector
from openpyxl.utils import get_column_letter

# 写入Excel文件数据


def writedata(path: str, database: DatabaseConnector,
              date: datetime.datetime, teacherName: str, teacherPhone: str, teacherEmail: str,
              teamLeaderName: str, teamLeaderPhone: str, teamLeaderEmail: str,
              workPlace: str, firstWage: float, secondWage: float, thirdWage: float, numForSubsidy: int):
    '''
    path: 文件保存路径
    database: MySQL连接
    date: 日期
    teacherName: 指导老师姓名
    teacherPhone: 指导老师电话
    teacherEmail: 指导老师邮箱
    teamLeaderName: 骨干姓名
    teamLeaderPhone: 骨干电话
    teamLeaderEmail: 骨干邮箱
    workPlace: 办公地点
    firstWage: 第一档工资
    secondWage: 第二档工资
    thirdWage: 第三档工资
    numForSubsidy: 建档立卡专设岗位
    '''
    # 创建excel，建立第一个表单Wage和第二个表单Info
    wb = openpyxl.Workbook()
    ws_sheet1 = wb.active
    ws_sheet1.title = 'Wage'
    ws_sheet2 = wb.create_sheet('Info')
    assert isinstance(ws_sheet2, Worksheet)
    # 读数据库
    database.execute("select * from AllMemberWageInfo;")
    results = database.fetchall()
    column_names = ["姓名", "学号", "部门名称", "岗位", "基本工资", "岗位备注",
                    "工资领取人姓名", "工资领取人学号", "银行卡号", "建档立卡", "个人备注"]  # 表头信息
    column_database_name = ["name", "student_id", "department_name", "job", "wage", "work_remark",
                            "application_name", "application_student_id", "application_bankcard", "subsidy_dossier", "wageinfo_remark"]  # 表头对应的数据库列名
    # Wage表单写入表头
    ws_sheet1.cell(row=1, column=1, value='{}年{}月勤工助学补助'.format(
        date.year, date.month))
    ws_sheet1.cell(row=2, column=1, value="指导老师姓名：{}        指导老师电话：{}         指导老师邮箱：{}".format(
        teacherName, teacherPhone, teacherEmail))
    ws_sheet1.cell(row=3, column=1, value='骨干姓名：{}         骨干电话：{}          骨干邮箱：{}         岗位总数：{}'.format(
        teamLeaderName, teamLeaderPhone, teamLeaderEmail, len(results), workPlace))
    ws_sheet1.cell(row=4, column=1, value='一档金额：{}元/月（人）           二档金额：{}元/月（人）           建档立卡专设岗位 {}（人）'.format(
        firstWage, secondWage, thirdWage, numForSubsidy))
    ws_sheet1.cell(row=5, column=1, value="序号")
    for conlum_index in range(2, len(column_names) + 2):
        ws_sheet1.cell(row=5, column=conlum_index,
                       value=column_names[conlum_index - 2])
    # Info表单写入表头
    ws_sheet2.cell(row=1, column=1, value='{}年{}月勤工助学补助'.format(
        date.year, date.month))
    ws_sheet2.cell(row=2, column=1, value="指导老师姓名：{}        指导老师电话：{}         指导老师邮箱：{}".format(
        teacherName, teacherPhone, teacherEmail))
    ws_sheet2.cell(row=3, column=1, value='骨干姓名：{}         骨干电话：{}          骨干邮箱：{}         岗位总数：{}         办公地点：{}								'.format(
        teamLeaderName, teamLeaderPhone, teamLeaderEmail, len(results), workPlace))
    ws_sheet2.cell(row=4, column=1, value='一档金额：{}元/月（人）   二档金额：{}元/月（人）            三档金额: {}元/月（人）            建档立卡专设岗位 {}（人）'.format(
        firstWage, secondWage, thirdWage, numForSubsidy))
    # 数据写入Info表单
    ws_sheet2.cell(row=5, column=1, value="序号")
    for conlum_index in range(2, len(column_names)+2):
        ws_sheet2.cell(row=5, column=conlum_index,
                       value=column_names[conlum_index-2])  # 写表头
    for row_index in range(1, len(results)+1):
        ws_sheet2.cell(row=row_index+5, column=1, value=row_index)
        for col_index in range(2, len(column_names)+2):
            ws_sheet2.cell(row=row_index+5, column=col_index,
                           value=results[row_index-1][column_database_name[col_index-2]])  # 写入表的内容
    ws_sheet2.cell(row=5+len(results)+1,
                   column=column_names.index('基本工资')+1, value='总金额')  # 写入‘总金额’
    ws_sheet2.cell(row=5+len(results)+1+2, column=1, value='1、各单位需完整填写表内各项，以便账号等出现问题可以及时取得联系。\n2、若该同学已离开该单位只需将该同学删除即可。新加入的同学请添加于最后，并用红色字体标出。补发的同学添加于新加入的同学之后，在发放月份栏注明“补发x月”。\n3、若该同学补助金额不为标准补助金额即对应的岗位设档金额时，请在备注说明原因。\n4、表的最后需计算总计金额。本表不能代替交至学工部的Word文档表。\n5、Excel表格及邮件主题名称为：部门+2019年 x 月）。\n6、请各单位注意核对同学（尤其是少数民族学生）的账号和姓名等信息，非建行新鸿支行卡不可用。\n7、若有建档立卡岗，请备注。\n8、第四行的设档信息是指申请并审核通过的设档信息，所以这一学期都是不会变化的。请勿错填成当月的成员信息。\n')
    # 保存文件
    wb.save(path)

# 设置Excel表单格式


def SetStyle(path: str):
    '''
    path: 文件路径
    '''
    wb = openpyxl.load_workbook(path)
    ws_sheet2 = wb['Info']  # 咱们先处理表二
    ws_sheet2.merge_cells('A1:L1')  # 前四行合并
    ws_sheet2.merge_cells('A2:L2')
    ws_sheet2.merge_cells('A3:L3')
    ws_sheet2.merge_cells('A4:L4')

    # 先设置第一行标题
    ws_sheet2.row_dimensions[1].height = 32  # 设置行高
    ws_sheet2.cell(row=1, column=1).font = Font(
        u'楷体', size=20, bold=True, color='000000')
    ws_sheet2.cell(row=1, column=1).alignment = Alignment(
        horizontal='center', vertical='center')

    # 2到4行的格式
    for R in range(2, 5):
        ws_sheet2.row_dimensions[R].height = 24
        ws_sheet2.cell(row=R, column=1).alignment = Alignment(
            horizontal='center', vertical='center')
        ws_sheet2.cell(row=R, column=1).font = Font(
            u'楷体', size=12, bold=True, color='000000')

    # 获得行数和列数
    nrows = ws_sheet2.max_row
    ncols = ws_sheet2.max_column

    for i in range(1, ncols+1):  # 第五行没合并就单独拿出来了
        ws_sheet2.cell(row=5, column=i).font = Font(
            u'楷体', size=12, bold=True, color='000000')

    trows = 0
    for r in range(5, nrows + 1):  # 数据格式
        for l in range(1, ncols + 1):
            ws_sheet2.row_dimensions[r].height = 21
            ws_sheet2.cell(row=r, column=l).alignment = Alignment(
                horizontal='center', vertical='center')
            ws_sheet2.cell(row=r, column=l).font = Font(
                u'楷体', size=12, color='000000')
            # 由于最下面总金额和下面的注意事项不放在边框内，所以真实行数不算他们
            if ws_sheet2.cell(row=r, column=1).value == None:
                trows = r - 2

    # 自适应列宽
    # 获取每一列的内容的最大宽度
    i = 0
    col_width = list()
    # 每列
    for col in ws_sheet2.columns:
        # 每行
        hanzi = 0  # 初始化判断是否有汉字的变量，0表示没汉字，1表示有汉字
        for j in range(5, len(col)-5):
            for c in str(col[j].value):  # 判断是否有汉字，汉字的列宽要增大
                if ('\u4e00' <= c <= '\u9fa5'):
                    hanzi = 1
            if j == 5:
                # 数组增加一个元素
                col_width.append(len(str(col[j].value)))
            else:
                # 获得每列中的内容的最大宽度
                if col_width[i] < len(str(col[j].value)):
                    col_width[i] = len(str(col[j].value))
        if hanzi == 1:
            col_width[i] = col_width[i]+6

        i = i + 1

    # 设置列宽

    # 根据列的数字返回字母
    for i in range(len(col_width)):
        col_letter = get_column_letter(i + 1)
        # 当宽度小于等于10，宽度设置为16
        if col_width[i] < 10 or col_width[i] == 10:
            ws_sheet2.column_dimensions[col_letter].width = 16
        # 只有当宽度大于10，才设置列宽
        elif col_width[i] > 10:
            ws_sheet2.column_dimensions[col_letter].width = 2*col_width[i] + 6

    # 设置边框
    for r in range(1, trows + 1):
        for l in range(1, ncols + 1):

            ws_sheet2.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                            right=Side(
                                                                border_style='thin', color='000000'),
                                                            top=Side(
                                                                border_style='thin', color='000000'),
                                                            bottom=Side(border_style='thin', color='000000'))
            # 粗实线设置
            if l == ncols and r != trows:
                ws_sheet2.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                                right=Side(
                                                                    border_style='medium', color='000000'),
                                                                top=Side(
                                                                    border_style='thin', color='000000'),
                                                                bottom=Side(border_style='thin', color='000000'))
            if r == trows and l != ncols:
                ws_sheet2.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                                right=Side(
                                                                    border_style='thin', color='000000'),
                                                                top=Side(
                                                                    border_style='thin', color='000000'),
                                                                bottom=Side(border_style='medium', color='000000'))
            if r == trows and l == ncols:
                ws_sheet2.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                                right=Side(
                                                                    border_style='medium', color='000000'),
                                                                top=Side(
                                                                    border_style='thin', color='000000'),
                                                                bottom=Side(border_style='medium', color='000000'))
    # 注意事项的格式设置
    ws_sheet2.merge_cells(start_row=trows+3, start_column=1,
                          end_row=trows+13, end_column=ncols)
    for i in range(trows+3, trows+11):
        for j in range(1, ncols + 1):
            ws_sheet2.cell(row=i, column=j).alignment = Alignment(
                wrapText=True)
            ws_sheet2.cell(row=i, column=j).font = Font(
                u'楷体', size=12,  color='000000')

    ws_sheet1 = wb['Wage']  # 咱再处理表一
    ws_sheet1.merge_cells('A1:L1')  # 前四行合并
    ws_sheet1.merge_cells('A2:L2')
    ws_sheet1.merge_cells('A3:L3')
    ws_sheet1.merge_cells('A4:L4')

    # 先设置第一行标题
    ws_sheet1.row_dimensions[1].height = 32  # 设置行高
    ws_sheet1.cell(row=1, column=1).font = Font(
        u'楷体', size=20, bold=True, color='000000')
    ws_sheet1.cell(row=1, column=1).alignment = Alignment(
        horizontal='center', vertical='center')

    # 2到4行的格式
    for R in range(2, 5):
        ws_sheet1.row_dimensions[R].height = 24
        ws_sheet1.cell(row=R, column=1).alignment = Alignment(
            horizontal='center', vertical='center')
        ws_sheet1.cell(row=R, column=1).font = Font(
            u'楷体', size=12, bold=True, color='000000')

    nrows = 40  # 空表，随便给点行数
    ncols = ws_sheet1.max_column
    trows = nrows-1
    # # 上面那个函数没写我就写在我这个部分了
    ws_sheet1.cell(row=nrows + 1, column=4 + 1, value='总金额')  # 写入‘总金额’
    ws_sheet1.cell(row=nrows + 1, column=4 +
                   1).font = Font(u'楷体', size=12, color='000000')
    ws_sheet1.cell(row=trows + 5, column=1,
                   value='1、各单位需完整填写表内各项，以便账号等出现问题可以及时取得联系。\n2、若该同学已离开该单位只需将该同学删除即可。新加入的同学请添加于最后，并用红色字体标出。补发的同学添加于新加入的同学之后，在发放月份栏注明“补发x月”。\n3、若该同学补助金额不为标准补助金额即对应的岗位设档金额时，请在备注说明原因。\n4、表的最后需计算总计金额。本表不能代替交至学工部的Word文档表。\n5、Excel表格及邮件主题名称为：部门+2019年 x 月）。\n6、请各单位注意核对同学（尤其是少数民族学生）的账号和姓名等信息，非建行新鸿支行卡不可用。\n7、若有建档立卡岗，请备注。\n8、第四行的设档信息是指申请并审核通过的设档信息，所以这一学期都是不会变化的。请勿错填成当月的成员信息。\n')

    for i in range(1, ncols + 1):  # 第五行没合并就单独拿出来了
        ws_sheet1.cell(row=5, column=i).font = Font(
            u'楷体', size=12, bold=True, color='000000')
        ws_sheet1.cell(row=5, column=i).alignment = Alignment(
            horizontal='center', vertical='center')
    for r in range(5, nrows + 1):  # 数据格式
        for l in range(1, ncols + 1):
            ws_sheet1.row_dimensions[r].height = 21
            ws_sheet1.cell(row=r, column=l).alignment = Alignment(
                horizontal='center', vertical='center')
            ws_sheet1.cell(row=r, column=l).font = Font(
                u'楷体', size=12, color='000000')

    for i in range(1, ncols + 1):  # 敷衍的设置一下列宽
        col_letter = get_column_letter(i + 1)
        ws_sheet1.column_dimensions[col_letter].width = 18
    # 设置边框
    for r in range(1, nrows + 1):
        for l in range(1, ncols + 1):

            ws_sheet1.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                            right=Side(
                                                                border_style='thin', color='000000'),
                                                            top=Side(
                                                                border_style='thin', color='000000'),
                                                            bottom=Side(border_style='thin', color='000000'))
            # 粗实线设置
            if l == ncols and r != nrows:
                ws_sheet1.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                                right=Side(
                                                                    border_style='medium', color='000000'),
                                                                top=Side(
                                                                    border_style='thin', color='000000'),
                                                                bottom=Side(border_style='thin', color='000000'))
            if r == nrows and l != ncols:
                ws_sheet1.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                                right=Side(
                                                                    border_style='thin', color='000000'),
                                                                top=Side(
                                                                    border_style='thin', color='000000'),
                                                                bottom=Side(border_style='medium', color='000000'))
            if r == nrows and l == ncols:
                ws_sheet1.cell(row=r, column=l).border = Border(left=Side(border_style='thin', color='000000'),
                                                                right=Side(
                                                                    border_style='medium', color='000000'),
                                                                top=Side(
                                                                    border_style='thin', color='000000'),
                                                                bottom=Side(border_style='medium', color='000000'))

    # 注意事项的格式设置
    ws_sheet1.merge_cells(start_row=trows + 5, start_column=1,
                          end_row=trows + 15, end_column=ncols)
    for i in range(trows + 3, trows + 11):
        for j in range(1, ncols + 1):
            ws_sheet1.cell(row=i, column=j).alignment = Alignment(
                wrapText=True)
            ws_sheet1.cell(row=i, column=j).font = Font(
                u'楷体', size=12,  color='000000')

    wb.save(path)


if __name__ == '__main__':
    '''
    一下是所需参数及其顺序和类型:
        path:str 文件保存路径目录
        date:datetime 日期
        teacherName:str 指导老师姓名
        teacherPhone:str 指导老师电话
        teacherEmail:str 指导老师邮箱
        teamLeaderName:str 骨干姓名
        teamLeaderPhone:str 骨干电话
        teamLeaderEmail:str 骨干邮箱
        workPlace:str 办公地点
        firstWage:float 第一档工资
        secondWage:float 第二档工资
        thirdWage:float 第三档工资
        numForSubsidy:int 建档立卡专设岗位
    '''
    try:
        input_len = len(sys.argv)-1
        if input_len == 13:
            session = DatabaseConnector()

            path = os.path.join(sys.argv[1], str(time.time()))
            os.makedirs(path, exist_ok=True)
            path = os.path.join(path, '财务报账表.xlsx')
            params = {
                'path': path,
                'database': session,
                'date': datetime.datetime.strptime(sys.argv[2], '%Y-%m-%d'),
                'teacherName': sys.argv[3],
                'teacherPhone': sys.argv[4],
                'teacherEmail': sys.argv[5],
                'teamLeaderName': sys.argv[6],
                'teamLeaderPhone': sys.argv[7],
                'teamLeaderEmail': sys.argv[8],
                'workPlace': sys.argv[9],
                'firstWage': float(sys.argv[10]),
                'secondWage': float(sys.argv[11]),
                'thirdWage': float(sys.argv[12]),
                'numForSubsidy': int(sys.argv[13])
            }
            writedata(**params)
            SetStyle(path)
        else:
            raise Exception('程序输入参数数量与需求不一致，获得了{}个参数'.format(input_len))
    except Exception as e:
        import json
        errors = {'ReturnCode': '417', 'ReturnString': '程序出错',
                  'ShowMessage': repr(e), 'Data': ''}
        print(json.dumps(errors, ensure_ascii=False))
    else:
        datetime.datetime.strptime('2013-2-3', '%Y-%m-%d')
        print(json.dumps({'ReturnCode': '200', 'ReturnString': '成功',
              'ShowMessage': '', 'Data': path}, ensure_ascii=False))
