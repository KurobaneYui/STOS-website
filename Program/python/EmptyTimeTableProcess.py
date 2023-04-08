# -*- coding:utf-8 -*-
import os
import sys
import json
import time
import datetime
import py7zr
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from Frame.python3.BaseComponents.DatabaseConnector import DatabaseConnector
from openpyxl.utils import get_column_letter

# 写入Excel文件数据


def writedata(path: str, database: DatabaseConnector):
    '''
    path: 文件保存路径
    database: MySQL连接
    '''
    # 每一个组的excel文件路径和名称列表
    excel_list = list()

    # 获取现场组每个组组号
    # 遍历每个组
    # # 对每个组建立一个excel文件
    # # 获取每个组的组员：姓名、学号、空课表数据
    # # 遍历获取到的组员信息
    # # # 对每个组员建立一个表单
    # # # 表单写入每个组员的空课信息
    # # 保存文件
    # 对所有的excel打包成7z文件
    # 删除中间excel

    database.execute("SELECT department_id, name as department_name \
                    FROM Department \
                    WHERE name like %s;",
                     data=("现场组%组",))
    departmentsInfo = database.fetchall()
    for one_department in departmentsInfo:
        wb = openpyxl.Workbook()
        ws_sheet_removable = wb.active
        ws_sheet_removable.title = 'removable'

        DBAffectedRows = database.execute("SELECT Work.student_id AS student_id, MemberBasic.name AS student_name, mon, tue, wed, thu, fri \
                                            FROM Work \
                                            LEFT JOIN MemberBasic ON Work.student_id = MemberBasic.student_id \
                                            LEFT JOIN EmptyTime ON EmptyTime.student_id = MemberBasic.student_id \
                                            WHERE job = 0 AND department_id = %s;",
                                          data=(one_department["department_id"],))
        studentsInfo = database.fetchall()

        if DBAffectedRows != 0:
            for one_student in studentsInfo:
                mon = one_student["mon"]
                tue = one_student["tue"]
                wed = one_student["wed"]
                thu = one_student["thu"]
                fri = one_student["fri"]

                ws = wb.create_sheet(one_student["student_name"])
                assert isinstance(ws, Worksheet)
                ws.append([one_student["student_id"],
                          "周一", "周二", "周三", "周四", "周五"])
                ws.append(["12节", mon[0], tue[0], wed[0], thu[0], fri[0]])
                ws.append(["34节", mon[1], tue[1], wed[1], thu[1], fri[1]])
                ws.append(["56节", mon[2], tue[2], wed[2], thu[2], fri[2]])
                ws.append(["78节", mon[3], tue[3], wed[3], thu[3], fri[3]])

            del wb['removable']

        # 保存为group1~group6，由于目前现场组1组的编号为5，所以用减4
        wb.save(os.path.join(
            os.path.split(path)[0],
            "group"+str(int(one_department["department_id"])-4)+".xlsx"
        ))
        excel_list.append(os.path.join(
            os.path.split(path)[0],
            "group"+str(int(one_department["department_id"])-4)+".xlsx"
        ))

    with py7zr.SevenZipFile(path, 'w') as archive:
        for i in excel_list:
            archive.write(i)

    for i in excel_list:
        os.remove(i)


if __name__ == '__main__':
    '''
    以下是所需参数及其顺序和类型:
        path:str 文件保存路径目录
    '''
    try:
        input_len = len(sys.argv)-1
        if input_len == 1:
            session = DatabaseConnector()
            session.startCursor()

            path = os.path.join(sys.argv[1], str(time.time()))
            os.makedirs(path, exist_ok=True)
            path = os.path.join(path, '空课表.7z')
            params = {
                'path': path,
                'database': session
            }
            writedata(**params)
        else:
            raise Exception('程序输入参数数量与需求不一致，获得了{}个参数'.format(input_len))
    except Exception as e:
        errors = {'ReturnCode': '417', 'ReturnString': '程序出错',
                  'ShowMessage': repr(e), 'Data': ''}
        print(json.dumps(errors, ensure_ascii=False))
    else:
        datetime.datetime.strptime('2013-2-3', '%Y-%m-%d')
        print(json.dumps({'ReturnCode': '200', 'ReturnString': '成功',
              'ShowMessage': '', 'Data': path}, ensure_ascii=False))
